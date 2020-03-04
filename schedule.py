# -- Schedule client program for air force -- #
# Fix these imports
import tkinter as tk
import openpyxl
import xlsxwriter
import datetime
import calendar
import os
import pickle # For data serialization

from tkinter import *
from PIL import Image
from PIL import ImageTk
from PIL import ImageOps
from openpyxl import load_workbook
from calendar import monthrange


# -- Function to add a number to a color -- #
def add_color(color, num):
    color_format = "0x" + color[1::]
    color_to_int = int(color_format, 16)
    new_int = color_to_int + num
    new_color = "#" + hex(new_int)[2:].upper()
    return new_color

# -- Function to resize and add a picture -- #
def add_pic(name, width, height):
    path = name
    open_path = Image.open(path)

    size = (width, height)
    fit_image = ImageOps.fit(open_path, size, Image.ANTIALIAS)

    down_bt_img = ImageTk.PhotoImage(fit_image)
    return down_bt_img

def onFrameConfigure(canvas):
    '''Reset the scroll region to encompass the inner frame'''
    canvas.configure(scrollregion=canvas.bbox("all"))

# Create the linux server and master field update program (Make it as a forever looped daemon)

# Color Scheme Note: https://www.color-hex.com/color-palette/72861
#https://www.sessions.edu/color-calculator/
# Probably will connect to a server running on Python that has the data field
# But right now we will not do that
# Could use object oriented

# == File Stuff Here == #
master_fields = open('fields/master.txt') # These are the master fields from the server
local_fields = open("fields/fields.txt", "w") # This is the user's local field sheet

# == Color Scheme == #
bg_color = "#E6F2FF" # Background color
bt_color = "#EEE6FF" # Button color

# == Fonts == #
bt_font = "Fixedsys"

# == Main Window == #
root = tk.Tk()

winx=1000 # Width of window
winy=400 # height of window

ws = root.winfo_screenwidth() # Width of the screen
hs = root.winfo_screenheight() # Height of the screen

# -- Calculate the x and y coordinates of the Tk root window
x = (ws/2) - (winx/2)
y = (hs/2) - (winy/2)

root.geometry("%dx%d+%d+%d" % (winx,winy, x, y)) # Window size

root.config(background=bg_color)
root.resizable(False, False)
root.title("Schedule Smart")

# == Image stuff == #
up_bt_img = add_pic("images/Up.png", winx//3, winy-110)
down_bt_img = add_pic("images/Down.png", winx//3, winy-110)

up_exist_bt_img = add_pic("images/Exist_Up.png", winx//3, winy-110)

air_sym_img = add_pic("images/air_symbol.png", winx//3, winy-10)


# == Data structures == #
v_que = {} # Used to hold the value and name of the checkboxes
    
# Make a frame for the other buttons
class Application(tk.Frame):
        # -- Application main variables -- #
        win_counter = 0 # Keeps track of number of sub windows
        grp_but_active = False # Used to keep track of the group button window
        
        # -- Initializes the application window -- #
        def __init__(self, master=None):
            super().__init__(master)
            self.master = master
            self.pack()
            self.create_widgets()

        # -- Creates the main button -- #
        def create_widgets(self):
                self.select_bt_frm = tk.Frame(root, bg=bg_color)
                self.select_bt_frm.pack(side="right")
                
                self.users_options = tk.Button(self.select_bt_frm, command=lambda: self.add_master(self.users_options, self.select_bt_frm), image = up_bt_img, bg = bg_color, height=winy, relief = FLAT, activebackground = bg_color, highlightthickness = 0)
                self.users_options.pack(side="right")
                
                self.users_options_exist = tk.Button(self.select_bt_frm, command=lambda: self.add_exist(self.users_options_exist, self.select_bt_frm), image = up_exist_bt_img, bg = bg_color, height=winy, relief = FLAT, activebackground = bg_color, highlightthickness = 0)
                self.users_options_exist.pack(side="right")
                
                self.air_logo = tk.Button(root, image = air_sym_img, bg = bg_color, height=winy, relief = FLAT, activebackground = bg_color, highlightthickness = 0)
                self.air_logo.pack(side="left")
                self.air_logo['state'] = "disabled"

        def add_exist(self, button, frame):
            button["image"] = down_bt_img
            
            root.after(200, lambda: self.exist_org(widget=button, container=frame))

        # -- Set up the results of the exist button -- #
        def exist_org(self, widget, container):
            widget.pack_forget()
            container.pack_forget()

            path, dirs, files = os.walk("schedules/").__next__() # Get a list of the schedules

            # -- Remove temporary files -- #
            for f in files:
                if(f.find("~$") is 0):
                    files.remove(f)

            # -- Build a frame for the contents of the exist options -- #
            self.exist_frm = tk.Frame(root, bg=bg_color)
            self.exist_frm.pack()

            variable = StringVar(root)
            variable.set(files[0]) # Default value
            
            w = OptionMenu(*(self.exist_frm, variable) + tuple(files))
            w.pack(side="top", pady=100)

            self.select_file = tk.Button(self.exist_frm, text="Select", width=winx//40, height=winy//20, font=(bt_font, 10), bg=bt_color, relief=FLAT)
            self.select_file.pack(side="top", pady=20)

            field_cells = pickle.load(open("serialized/schedule.txt", "rb"))
            print(field_cells)
            
        # -- Add master list button -- #
        def add_master(self, button, frame):
            button["image"] = down_bt_img
            
            root.after(200, lambda: self.img_org(widget=button, container=frame))
            
        # -- Adds the checkboxes -- #
        
        # Change the name of this function
        def img_org(self, widget, container):
                widget.pack_forget()
                container.pack_forget()

                # -- This canvas is used to hold the checkboxes in another window inside the app -- #
                canvas = tk.Canvas(root, borderwidth=0, background=bg_color, bd=0, highlightthickness=0)
                frame = Frame(canvas, width=winx//5, height=winy*2.5, bg=bg_color)
                vsb = tk.Scrollbar(root, orient="horizontal", command=canvas.xview, background=bg_color)
                canvas.configure(xscrollcommand=vsb.set)

                canvas.pack(side="top", fill="both", expand=True)
                vsb.pack(side="top", fill="x")
                canvas.create_window((4,4), window=frame, anchor="nw")

                frame.bind("<Configure>", lambda event, canvas=canvas: onFrameConfigure(canvas))
                
                global v_que
                
                x = 0
                y = 0
                
                count = 0
                n = "B"+str(count)
                val = BooleanVar(name=n)
                
                line = master_fields.readline()
                
                v_que[line] = val
                
                li = self.add_chk_but(line, x, y, val, bg_color, frame)
                
                while(line is not ''):
                        li = self.add_chk_but(line, x, y, val, bg_color, frame) # Calls checkbox list function
                        
                        line = master_fields.readline()
                        x+=1
                        if(x % 5 == 0):
                                y+=1
                                x=0
                        
                        count+=1
                        n="B"+str(count)
                        val = BooleanVar(name=n)
                        v_que[line] = val
                        
                self.spawn_frm = Frame(root, width=winx//5, height=winy*2.5, bg=bg_color) # Used to keep track of the checkbox spawning buttons
                self.spawn_frm.pack(side="bottom")
                
                self.spawn_bt = tk.Button(self.spawn_frm, text="Select Fields", font=(bt_font, 10), command=lambda: self.spawn_list(self.spawn_frm), width=winx//30, height=winy//4, bg=bt_color, relief = FLAT)
                self.spawn_bt.pack(side="left", padx=3, pady=9)

                self.spawn_grp_bt = tk.Button(self.spawn_frm, text="Group", font=(bt_font, 10), command=lambda:self.add_group(), width=winx//30, height=winy//4, bg=bt_color, relief = FLAT)
                self.spawn_grp_bt.pack(side="left", padx=3, pady=9)
            
        # == Spawn the list based off the fields chosen == #
        def spawn_list(self, frame):
                global local_fields # Local fields list
                
                frame.pack_forget() # Hide the last button
                count=0 # Counts how many variables are selected

                # -- Fill the local list with variables -- #
                
                for key, value in v_que.items():
                        if(value.get() is True):
                                local_fields.write(key)
                                count+=1
                        
                local_fields.close() # Close to save the changes
                
                self.create_sheet_bt = tk.Button(root, text="Create Sheet", font=(30), command=lambda: self.create_sheet(self.create_sheet_bt, count), width=winx//10, height=winy//4, bg=bt_color, relief = FLAT)
                self.create_sheet_bt.pack(padx=3, pady=9)

        # == Creates the worksheet from the local fields list == #
        def create_sheet(self, button, count):
                button.pack_forget() # Forgets that button that was clicked

                # -- Calender data -- #
                d = datetime.datetime.now() # Gets a data structre of the current time
                month = d.month
                year = d.year
                lastday = monthrange(year, month)[1] # Gets the number of the month
                
                # -- Add new buttons -- #
                self.autofill_bt = tk.Button(root, command=lambda:self.add_group(), text="Group", font=(bt_font, 30), width=winx//120, height=winy//16, bg=bt_color, relief = FLAT)
                self.autofill_bt.pack(side="left", padx=10, pady=3)

                self.edit_bt = tk.Button(root, text="Save", font=(bt_font, 30), width=winx//120, height=winy//16, bg=bt_color, relief = FLAT)
                self.edit_bt.pack(side="left", padx=10, pady=3)

                self.update_bt = tk.Button(root, text="Update", font=(bt_font, 30), width=winx//120, height=winy//16, bg=bt_color, relief = FLAT)
                self.update_bt.pack(side="left", padx=10, pady=3)

                # -- xlsxwriter functions (Set the workbook up) -- ##
                workbook = xlsxwriter.Workbook("schedules/schedule.xlsx") # Creates a new workbook called schedule
                worksheet = workbook.add_worksheet("Schedule") # Creates a new worksheet called Schedule

                worksheet.freeze_panes(0, 1) # Freezes the leftmost pane so that you can scroll freely without losing sight of it

                # -- Hides the unused row cells -- #
                cell_num = "A" + str(count)
                worksheet.write(cell_num, 'Used')
                worksheet.set_default_row(hide_unused_rows=True)

                # -- Hides the unused column cells -- #
                cell_num = "A" + chr(lastday+65-26) + ":XFD"
                worksheet.set_column(cell_num, None, None, {'hidden': True})

                cell_format = workbook.add_format()
                cell_format.set_pattern(1)
                
                cell_color = "#84F2CF"
                
                cell_format.set_bg_color(cell_color)

                for i in range(2, lastday+2):
                    if((i+64)>90):
                        num = 'A'
                        num = num + chr(i+64-26) + str(1)
                    else:
                        num = chr(i+64) + str(1)
                    
                    worksheet.write(num, "", cell_format)

                    cell_format = workbook.add_format()
                    cell_format.set_pattern(1)
                        
                    cell_color = add_color(cell_color, 21)
                    cell_format.set_bg_color(cell_color)
                    
                cell_color = "#FFF2CF"
                
                cell_format.set_bg_color(cell_color)

                for i in range(2, count+2):
                    num = 'A' + str(i)
                    worksheet.write(num, "", cell_format)
                    
                    cell_format = workbook.add_format()
                    cell_format.set_pattern(1)
                    
                    cell_color = add_color(cell_color, 50)
                    cell_format.set_bg_color(cell_color)

                workbook.close()

                # -- openpyxl (Write in the cell values) -- #
                wb = load_workbook("schedules/schedule.xlsx")
                sheet = wb["Schedule"]
                wb.active = sheet
                
                field_list = open("fields/fields.txt", "r")
                
                for i in range(1, lastday+1):
                        cur = sheet.cell(row=1, column=i+1)
                        cur.value = "%s %d" % (calendar.month_name[month], i)
                        _cell = sheet.cell(row=1, column=i+1)

                        if((i+65)>90):
                                num = 'A'
                                num = num + chr(i+65-26)
                        else:
                                num = chr(i+65)
                        
                        sheet.column_dimensions[num].width = len(cur.value) + 1
                        
                cur = sheet.cell(row=1, column=1)
                cur.value = "Key"
                cur.font = '34 pt bold'

                max_val = []
                field_cells = [] # A list of the fields used, for later use in existing schedule options
                
                for i in range(2, count+2):
                        cur = sheet.cell(row=i, column=1)
                        cur.value = field_list.readline()
                        max_val += [len(cur.value)]
                        cur.font = '24 pt bold'

                        field_cells += [[i, 1]] # Add cordinates of cell

                # -- Add unique identifier or name to schedule.txt. Maybe maintain a file that list serialized data for schedules -- #
                serial_field_cells = pickle.dump(field_cells, open("serialized/schedule.txt", "wb"))
                
                cell_width = max(max_val)
                sheet.column_dimensions['A'].width = cell_width+1
                
                field_list.close()
                #sheet.row_dimensions[5].hidden = False
                wb.save('schedules/schedule.xlsx')
                wb.close()

        # -- Add group options -- #
        def add_group(self):
            if self.grp_but_active is True:
                return
            else:
                self.grp_but_active = True
            
            self.win_counter += 1 # Increment the window counter

            t = tk.Toplevel(self) # Spawn a new window
            
            groupx = winx//2
            groupy = winy//2 + 160

            ws = root.winfo_screenwidth() # Width of the screen
            hs = root.winfo_screenheight() # Height of the screen

            x = (ws/2) - (groupx/2) - (winx/2) # Move to the left of the main app
            y = (hs/2) - (groupy/2)

            # -- Colors for group window -- #
            group_bg_color = "#b3d9ff"
            group_bt_color = "#cbb3ff"
            
            t.geometry("%dx%d+%d+%d" % (groupx, groupy, x, y))
            t.config(background=group_bg_color)
            t.resizable(False, False)

            t.protocol("WM_DELETE_WINDOW", lambda: self.on_close(window=t)) # To handle when the window is exited
            
            t.title("Group Select")

            # -- Spawns the group frame for checkboxes -- #
            grp_canvas = tk.Canvas(t, borderwidth=0, background=group_bg_color, bd=0, highlightthickness=0)
            grp_frame = Frame(grp_canvas, width=groupx, height=groupy, bg=group_bg_color)
            grp_vsb = tk.Scrollbar(t, orient="horizontal", command=grp_canvas.xview, background=group_bg_color)
            grp_canvas.configure(xscrollcommand=grp_vsb.set)

            grp_canvas.pack(side="top", fill="both", expand=True)
            grp_vsb.pack(side="top", fill="x")
            grp_canvas.create_window((4,4), window=grp_frame, anchor="nw")

            grp_frame.bind("<Configure>", lambda event, canvas=grp_canvas: onFrameConfigure(grp_canvas))
                
            # -- Adds the group options -- #
            grp_que = {} # Queue to hold the vales of the group checkboxes
            
            x = 0
            y = 0
            
            count = 0
            n = "C" + str(count)
            val = BooleanVar(name=n)

            path, dirs, files = os.walk("groups/").__next__() # Get a list of the files in group
            file_count = len(files) # The number of files in the group folder
            
            line = files[0].strip(".txt") # Gets the group name without the file extension
            li = self.add_chk_but(line, x, y, val, group_bg_color, grp_frame)
            grp_que[line] = val
            
            for i in range(1, file_count):
                x+=1
                if(x%5 == 0):
                    y+=1
                    x=0

                count+=1
                n="C" + str(count)
                val = BooleanVar(name=n)
                line = files[i].strip(".txt") # Gets the group name without the file extension
                
                grp_que[line] = val

                li = self.add_chk_but(line, x, y, val, group_bg_color, grp_frame)

            self.select_frm = tk.Frame(t, bg=group_bg_color)
            self.select_frm.pack()
            
            self.select_grp_but = tk.Button(self.select_frm, text="Select", command=lambda:self.update_fields_grp(groups=grp_que, select=True), font=(bt_font, 10), width=groupx//20, height=groupy//2, bg=group_bt_color, relief=FLAT)
            self.select_grp_but.pack(side="left", padx=2, pady=2)

            self.select_grp_but = tk.Button(self.select_frm, text="Deselect", command=lambda:self.update_fields_grp(groups=grp_que, select=False), font=(bt_font, 10), width=groupx//20, height=groupy//2, bg=group_bt_color, relief=FLAT)
            self.select_grp_but.pack(side="left", padx=2, pady=2)

        # -- Updates the current fields using the checkbox fields group selection -- #
        def update_fields_grp(self, groups, select):
            global v_que
            
            file_list = [] # List of chosen group files
            fields_list = set() # List of fields selected from groups, uses a set to remove repeats
            
            for key, value in groups.items():
                if value.get() is True:
                    file_list+=["groups/"+key+".txt"] # If selected then add that file name

            for name in file_list:
                try:
                    f = open(name, 'r') # Open the selected group file
                except Exception as e:
                    print("Could not open %s" % name)
                    
                field = f.readline().rstrip() # Adds the first field, removes the newline characters
                fields_list.add(field)
                
                while(field is not ''):
                    field = f.readline().rstrip()
                    fields_list.add(field)

                f.close() # Close the selected group

            for key, value in v_que.items():
                if key.rstrip() in fields_list:
                    # -- Select fields based off which button was pressed -- #
                    if(select is True and key.rstrip() is not ''):
                        value.set(True)
                    elif(select is False and key.rstrip() is not ''):
                        value.set(False)
                

            
        def on_close(self, window):   
            self.grp_but_active = False

            window.destroy()
            
            print("Closing window")

        # -- Creates check buttons in a frame-- #
        def add_chk_but(self, name, x, y, val, color, frame):
                tk.Checkbutton(frame, text=name, bg=color, variable=val).grid(row=x, column=y, padx=5, pady=5, sticky=W)

# -- Main loop -- #
app = Application(master=root)
app.mainloop()

master_fields.close()
local_fields.close()


